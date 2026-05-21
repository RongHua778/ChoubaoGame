# -*- coding: utf-8 -*-
"""Read Data/carddesign.xlsx（sheet「全卡表」/「负面卡表」/「特殊卡表」）写入 Data/cards.json。

列「牌名」或「牌名列」（若都存在则优先前者在 NAME_COLUMNS 中的顺序——当前为先「牌名列」再「牌名」）
写入各卡的 card.displayName，供卡组表按中文解析；商店是否上架仍由运行时与「是否解锁」列控制。"""
from __future__ import annotations

import json
import math
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("请先安装依赖: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "Data" / "carddesign.xlsx"
OUTPUT_PATH = ROOT / "Data" / "cards.json"
SHEET_FULL_TABLE = "全卡表"
SHEET_NEGATIVE_TABLE = "负面卡表"
SHEET_SPECIAL_TABLE = "特殊卡表"
UNLOCK_COLUMN = "是否解锁"
# 「牌名列」或与数据表一致的「牌名」，用于卡组表按中文名解析与 UI 展示
NAME_COLUMNS = ("牌名列", "牌名")

EFFECT_BY_CARD_ID = {
    4: "ambush",
    5: "copy",
    6: "bruiser",
    7: "vanguard",
    8: "imp",
    9: "divineLight",
    10: "shield",
    11: "chieftain",
    12: "steal",
    13: "bluff",
    14: "intimidate",
    15: "leech",
    16: "flood",
    17: "pollute",
    18: "menace",
    19: "incite",
    20: "wasteland",
    21: "poisonSource",
    22: "leakSecrets",
    23: "swap",
    24: "support",
    25: "amplify",
    26: "legion",
    27: "strengthen",
    28: "nimble",
    29: "frontrunner",
    30: "hook",
    31: "defuse",
    32: "chopsticks",
    33: "detonate",
    34: "threaten",
    35: "calm",
    36: "bigFish",
    37: "bait",
    38: "salvage",
    39: "rushOrder",
    40: "badReview",
    41: "delivery",
    42: "badCoin",
    43: "fakeGoods",
}

EFFECT_BY_NAME = {
    "偷窃": "steal",
    "虚张": "bluff",
    "恐吓": "intimidate",
    "陷阱": "ambush",
    "埋伏": "ambush",
    "复制": "copy",
    "莽夫": "bruiser",
    "先锋": "vanguard",
    "小鬼": "imp",
    "圣光": "divineLight",
    "回收": "shield",
    "回滚": "shield",
    "首领": "chieftain",
    "魅魔": "leech",
    "掀桌": "flood",
    "污染": "pollute",
    "贷款": "menace",
    "勾引": "incite",
    "怂恿": "incite",
    "废土": "wasteland",
    "毒源": "poisonSource",
    "泄密": "leakSecrets",
    "引爆": "detonate",
    "自爆": "detonate",
    "威胁": "threaten",
    "冷静": "calm",
    "大鱼": "bigFish",
    "鱼饵": "bait",
    "打捞": "salvage",
    "抢单": "rushOrder",
    "差评": "badReview",
    "外卖": "delivery",
    "劣币": "badCoin",
    "假货": "fakeGoods",
    "筷子": "chopsticks",
    "断点": "defuse",
    "腐蚀": "corrode",
    "封禁": "lockdown",
    "火药": "gunpowder",
}


def _cell_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()


def _cell_int(val, default: int = 0) -> int:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return default
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _headers_include_unlock(headers: list[str]) -> bool:
    return UNLOCK_COLUMN in headers


def _row_unlocked_flag(row: dict, headers: list[str]) -> bool:
    """从「是否解锁」列解析布尔标签；无该列或未填写视为 True（全部上架）。"""
    if not _headers_include_unlock(headers):
        return True

    val = row.get(UNLOCK_COLUMN)
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)) and not (isinstance(val, float) and math.isnan(val)):
        return float(val) != 0

    s = _cell_str(val).lower()
    if not s:
        # 新加列后旧行未填：默认可导出，避免一次加列清空整张卡表
        return True

    unlocked_tokens = (
        "是",
        "y",
        "yes",
        "true",
        "t",
        "1",
        "解锁",
        "已解锁",
        "√",
        "✓",
        "开",
        "启用",
        "on",
    )
    locked_tokens = (
        "否",
        "n",
        "no",
        "false",
        "f",
        "0",
        "未解锁",
        "锁",
        "锁定",
        "关",
        "禁用",
        "off",
    )
    if s in unlocked_tokens:
        return True
    if s in locked_tokens:
        return False

    # 其它写法：保守视为未解锁（商店不出售，仍会写入 cards.json）
    return False


def row_to_entry(row: dict, unlocked: bool, negative: bool = False, special: bool = False) -> dict | None:
    rid = row.get("ID")
    if rid is None or (isinstance(rid, float) and math.isnan(rid)):
        return None

    entry_id = _cell_int(rid, -1)
    if entry_id < 0:
        return None

    shop_price = _cell_int(row.get("商店价格"), 4)
    base_score = _cell_int(row.get("基础分"), 0)
    effect_name = _cell_str(row.get("效果名"))
    effect_desc = _cell_str(row.get("效果"))
    if entry_id == 33 and effect_name == "引爆":
        effect_name = "自爆"

    if "炸弹" in effect_name:
        card = {"type": "bomb", "value": 0}
    else:
        card = {"type": "score", "value": base_score}
        effect_key = (None if negative or special else EFFECT_BY_CARD_ID.get(entry_id)) or EFFECT_BY_NAME.get(effect_name)
        if effect_key:
            card["effect"] = effect_key
        if negative:
            card["negativeCard"] = True
        if special:
            card["specialCard"] = True

    name_col = ""
    for col in NAME_COLUMNS:
        v = row.get(col)
        if v is None or (isinstance(v, float) and math.isnan(v)):
            continue
        s = _cell_str(v)
        if s:
            name_col = s
            break

    if name_col:
        if entry_id == 33 and name_col == "引爆":
            name_col = "自爆"
        card["displayName"] = name_col
    elif card["type"] == "bomb":
        card["displayName"] = effect_name if effect_name else "炸弹"
    elif effect_name:
        card["displayName"] = effect_name
    else:
        card["displayName"] = "点数 {}".format(card["value"])

    return {
        "id": entry_id,
        "negative": negative,
        "special": special,
        "unlocked": unlocked,
        "shopPrice": shop_price,
        "baseScore": base_score,
        "effectName": effect_name,
        "effectDescription": effect_desc,
        "card": card,
    }


def read_full_table(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_FULL_TABLE not in wb.sheetnames:
        names = list(wb.sheetnames)
        wb.close()
        raise KeyError("未找到工作表「{}」，当前工作表: {}".format(SHEET_FULL_TABLE, names))

    ws = wb[SHEET_FULL_TABLE]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter)
    headers = [_cell_str(h) for h in header_row]

    out: list[dict] = []
    for raw in rows_iter:
        if not raw or all(c is None or (isinstance(c, float) and math.isnan(c)) for c in raw):
            continue
        row = dict(zip(headers, raw))
        unlocked = _row_unlocked_flag(row, headers)
        entry = row_to_entry(row, unlocked)
        if entry:
            out.append(entry)

    wb.close()
    unlocked_count = sum(1 for e in out if e.get("unlocked"))
    if _headers_include_unlock(headers):
        print("其中 {} / {} 张标记为已解锁（商店池）".format(unlocked_count, len(out)))
    out.sort(key=lambda e: e["id"])
    return out


def read_negative_table(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NEGATIVE_TABLE not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[SHEET_NEGATIVE_TABLE]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []
    headers = [_cell_str(h) for h in header_row]

    out: list[dict] = []
    for raw in rows_iter:
        if not raw or all(c is None or (isinstance(c, float) and math.isnan(c)) for c in raw):
            continue
        row = dict(zip(headers, raw))
        entry = row_to_entry(row, False, negative=True)
        if entry:
            out.append(entry)

    wb.close()
    out.sort(key=lambda e: e["id"])
    print("读取负面卡 {} 张".format(len(out)))
    return out


def read_special_table(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_SPECIAL_TABLE not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[SHEET_SPECIAL_TABLE]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []
    headers = [_cell_str(h) for h in header_row]

    out: list[dict] = []
    for raw in rows_iter:
        if not raw or all(c is None or (isinstance(c, float) and math.isnan(c)) for c in raw):
            continue
        row = dict(zip(headers, raw))
        entry = row_to_entry(row, False, special=True)
        if entry:
            out.append(entry)

    wb.close()
    out.sort(key=lambda e: e["id"])
    print("读取特殊卡 {} 张".format(len(out)))
    return out


def main() -> None:
    if not XLSX_PATH.is_file():
        print("找不到表格文件: {}".format(XLSX_PATH), file=sys.stderr)
        sys.exit(1)

    cards = read_full_table(XLSX_PATH)
    negative_cards = read_negative_table(XLSX_PATH)
    special_cards = read_special_table(XLSX_PATH)
    payload = {
        "version": 1,
        "source": "carddesign.xlsx",
        "sheet": SHEET_FULL_TABLE,
        "negativeSheet": SHEET_NEGATIVE_TABLE,
        "specialSheet": SHEET_SPECIAL_TABLE,
        "cards": cards,
        "negativeCards": negative_cards,
        "specialCards": special_cards,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print("已写入 {} (共 {} 张定义)".format(OUTPUT_PATH, len(cards)))


if __name__ == "__main__":
    main()
