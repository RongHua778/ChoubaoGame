# -*- coding: utf-8 -*-
"""Read Data/round1decks.xlsx and overwrite Data/Decks/Rounds/round_1/*.json opponents.

先于本脚本运行 export_cards.py。卡组「卡组」列可写 catalog 数字 ID、全卡表中牌名列/牌名列出的唯一中文名（与 cards.json 里 card.displayName 一致）；「炸弹」或 0 → catalog id 0。导出的 JSON 中 cardIds 一律为数字。"""
from __future__ import annotations

import json
import math
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("请先安装依赖: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
CARDS_JSON_PATH = ROOT / "Data" / "cards.json"
ROUND1_XLSX_PATH = ROOT / "Data" / "round1decks.xlsx"
OUTPUT_DIR = ROOT / "Data" / "Decks" / "Rounds" / "round_1"

REQUIRED_COLUMNS = ("轮次", "定位", "卡组名", "风格", "卡组")

POSITION_SLOTS: dict[str, tuple[str, ...]] = {
    "切磋": ("normal_a", "normal_b", "normal_c"),
    "强手": ("elite_a", "elite_b"),
    "庄家": ("boss",),
}

TOKEN_SPLIT_RE = re.compile(r"[，、,；;]+")

AI_STYLE_BY_LABEL = {
    "保守": "conservative",
    "胆小": "conservative",
    "膽小": "conservative",
    "谨慎": "conservative",
    "穩健": "balanced",
    "稳健": "balanced",
    "灵活": "balanced",
    "靈活": "balanced",
    "平衡": "balanced",
    "普通": "balanced",
    "激进": "aggressive",
    "激進": "aggressive",
    "暴躁": "aggressive",
    "莽撞": "aggressive",
}


def normalize_ai_style(style: str) -> str:
    s = _cell_str(style)
    if not s:
        return "balanced"
    lower = s.lower()
    if lower in ("conservative", "balanced", "aggressive"):
        return lower
    for key, value in AI_STYLE_BY_LABEL.items():
        if key in s:
            return value
    return "balanced"


def _cell_int(val, default: int = 0) -> int:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return default
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _cell_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()


def load_cards_catalog(path: Path) -> dict:
    if not path.is_file():
        print("请先运行 scripts/export_cards.py 生成 {}".format(path), file=sys.stderr)
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def build_name_to_catalog_id(payload: dict) -> dict[str, int]:
    """仅以每条目录 card.displayName（全卡表牌名）建唯一映射；「炸弹」→ 0 为约定写法。"""
    if not payload or not isinstance(payload.get("cards"), list):
        raise ValueError("cards.json 格式无效：缺少 cards 数组")

    out: dict[str, int] = {}

    def reg(key: str, cid: int) -> None:
        if not key:
            return
        prev = out.get(key)
        if prev is None:
            out[key] = cid
        elif prev != cid:
            print('警告：牌名「{}」重复对应 id {} 与 {}，保留先出现的 {}'.format(key, prev, cid, prev), file=sys.stderr)

    for entry in payload["cards"]:
        cid = entry.get("id")
        if not isinstance(cid, int):
            continue
        card = entry.get("card") or {}
        if isinstance(card, dict):
            dn = _cell_str(card.get("displayName"))
            reg(dn, cid)

    out["炸弹"] = 0
    return out


def split_deck_tokens(raw: str) -> list[str]:
    s = _cell_str(raw)
    if not s:
        return []
    parts = [p.strip() for p in TOKEN_SPLIT_RE.split(s)]
    return [p for p in parts if p]


def lookup_card_ids(tokens: list[str], name_map: dict[str, int]) -> tuple[list[int], list[str]]:
    missing: list[str] = []
    ids: list[int] = []
    for tok in tokens:
        if tok.isdigit():
            ids.append(int(tok))
            continue
        cid = name_map.get(tok)
        if cid is None:
            missing.append(tok)
        else:
            ids.append(cid)
    return ids, missing


def read_round1_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header_row = next(it)
        headers = [_cell_str(h) for h in header_row]
        for col in REQUIRED_COLUMNS:
            if col not in headers:
                wb.close()
                raise KeyError("round1decks.xlsx 表头缺少「{}」，当前列为: {}".format(col, headers))

        rows: list[dict] = []
        for raw in it:
            if not raw or all(c is None or (isinstance(c, float) and math.isnan(c)) for c in raw):
                continue
            row = dict(zip(headers, raw))
            if _cell_int(row.get("轮次"), -1) != 1:
                continue
            rows.append(row)
    finally:
        wb.close()
    return rows


def stem_for_position(pos: str, index: int) -> str:
    slots = POSITION_SLOTS.get(pos)
    if not slots:
        raise ValueError('未知的「定位」: "{}"（须为 {}）'.format(pos, " / ".join(POSITION_SLOTS.keys())))
    if index < 0 or index >= len(slots):
        raise ValueError('「{}」卡组数量溢出：至多 {} 套，表格中出现第 {} 套'.format(pos, len(slots), index + 1))
    return slots[index]


def load_existing_opponent_lines(path: Path) -> dict | None:
    if not path.is_file():
        return None
    try:
        old = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    ol = old.get("opponentLines")
    return ol if isinstance(ol, dict) else None


def write_slot_json(stem: str, deck_display_name: str, pos_label: str, style: str, card_ids: list[int]) -> None:
    out_path = OUTPUT_DIR / "{}.json".format(stem)
    deck_id = "round_1_{}".format(stem)
    old_lines = load_existing_opponent_lines(out_path)
    opponent_lines = old_lines or {}
    desc = "第1轮｜{}｜风格：{}".format(pos_label, style if style.strip() else "—").strip()
    ai_style = normalize_ai_style(style)

    payload = {
        "version": 1,
        "deckId": deck_id,
        "name": deck_display_name,
        "description": desc.strip(),
        "aiStyle": ai_style,
        "cardIds": card_ids,
        "opponentLines": opponent_lines,
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("已写入 {}".format(out_path.relative_to(ROOT)))


def main() -> None:
    if not ROUND1_XLSX_PATH.is_file():
        print("找不到表格: {}".format(ROUND1_XLSX_PATH), file=sys.stderr)
        sys.exit(1)

    catalog = load_cards_catalog(CARDS_JSON_PATH)
    name_map = build_name_to_catalog_id(catalog)

    rows = read_round1_rows(ROUND1_XLSX_PATH)
    if not rows:
        print("round1decks.xlsx 中没有轮次=1 的数据行", file=sys.stderr)
        sys.exit(1)

    counters: dict[str, int] = {k: 0 for k in POSITION_SLOTS}

    for row in rows:
        pos = _cell_str(row.get("定位"))
        if not pos:
            print("跳过：「定位」为空", file=sys.stderr)
            continue
        idx = counters.get(pos, -1)
        if idx < 0:
            print('跳过：未知定位「{}」'.format(pos), file=sys.stderr)
            continue

        try:
            stem = stem_for_position(pos, idx)
        except ValueError as e:
            print(e, file=sys.stderr)
            sys.exit(1)

        counters[pos] = idx + 1

        deck_title = _cell_str(row.get("卡组名"))
        style = _cell_str(row.get("风格"))
        tokens = split_deck_tokens(_cell_str(row.get("卡组")))
        card_ids, missing = lookup_card_ids(tokens, name_map)
        if missing:
            print('卡组「{}」中有未在 cards.json 中解析的名称: {}'.format(deck_title or stem, missing), file=sys.stderr)
            print("（请先在 carddesign 全卡表中配置「牌名/牌名列」并运行 export_cards.py）", file=sys.stderr)
            sys.exit(1)

        write_slot_json(stem, deck_title if deck_title else stem, pos, style, card_ids)

    expected = {"切磋": 3, "强手": 2, "庄家": 1}
    for lab, exp in expected.items():
        got = counters.get(lab, 0)
        if got != exp:
            print(
                '提示：第一轮期望 {} {} 套，实际表中为 {} 套（游戏仍加载 JSON，但未按设计数量写入）'.format(exp, lab, got),
                file=sys.stderr,
            )


if __name__ == "__main__":
    main()
